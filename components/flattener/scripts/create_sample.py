"""
Create a sample .xlsm file with data for testing the flattener.

Note: openpyxl cannot create VBA macros programmatically.
This script creates the structure, but you'll need to add VBA manually in Excel.

Alternatively, use the provided sample.xlsm file if available.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
from pathlib import Path


def create_sample_workbook():
    """
    Create a sample workbook with various Excel features.
    """
    print("Creating sample workbook...")

    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # ===== Sheet 1: Sales Data =====
    ws1 = wb.create_sheet('Sales Data', 0)

    # Headers
    headers = ['Date', 'Product', 'Quantity', 'Price', 'Total']
    for col, header in enumerate(headers, start=1):
        cell = ws1.cell(1, col, header)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # Sample data
    data = [
        [datetime(2025, 1, 1), 'Widget A', 10, 25.50, '=C2*D2'],
        [datetime(2025, 1, 2), 'Widget B', 5, 42.00, '=C3*D3'],
        [datetime(2025, 1, 3), 'Widget A', 15, 25.50, '=C4*D4'],
        [datetime(2025, 1, 4), 'Widget C', 8, 15.75, '=C5*D5'],
        [datetime(2025, 1, 5), 'Widget B', 12, 42.00, '=C6*D6'],
    ]

    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws1.cell(row_idx, col_idx, value)

    # Format dates
    for row in range(2, 7):
        ws1.cell(row, 1).number_format = 'YYYY-MM-DD'

    # Format prices
    for row in range(2, 7):
        ws1.cell(row, 4).number_format = '£#,##0.00'
        ws1.cell(row, 5).number_format = '£#,##0.00'

    # Totals row
    ws1.cell(7, 1, 'Total:')
    ws1.cell(7, 1).font = Font(bold=True)
    ws1.cell(7, 3, '=SUM(C2:C6)')
    ws1.cell(7, 5, '=SUM(E2:E6)')
    ws1.cell(7, 5).font = Font(bold=True)
    ws1.cell(7, 5).number_format = '£#,##0.00'

    # Create Excel table
    table = Table(displayName='SalesTable', ref='A1:E6')
    style = TableStyleInfo(
        name='TableStyleMedium2',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws1.add_table(table)

    # Add chart
    chart = LineChart()
    chart.title = 'Sales Trend'
    chart.x_axis.title = 'Date'
    chart.y_axis.title = 'Total Sales'

    # Chart data
    data_ref = Reference(ws1, min_col=5, min_row=1, max_row=6)
    cats_ref = Reference(ws1, min_col=2, min_row=2, max_row=6)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    ws1.add_chart(chart, 'G2')

    # ===== Sheet 2: Summary =====
    ws2 = wb.create_sheet('Summary', 1)

    ws2.cell(1, 1, 'Summary Report')
    ws2.cell(1, 1).font = Font(bold=True, size=14)

    ws2.cell(3, 1, 'Total Revenue:')
    ws2.cell(3, 2, "='Sales Data'!E7")
    ws2.cell(3, 2).number_format = '£#,##0.00'

    ws2.cell(4, 1, 'Total Items Sold:')
    ws2.cell(4, 2, "='Sales Data'!C7")

    ws2.cell(6, 1, 'Average Sale:')
    ws2.cell(6, 2, '=B3/B4')
    ws2.cell(6, 2).number_format = '£#,##0.00'

    # ===== Sheet 3: Hidden Sheet =====
    ws3 = wb.create_sheet('Configuration', 2)
    ws3.cell(1, 1, 'Tax Rate:')
    ws3.cell(1, 2, 0.20)
    ws3.cell(2, 1, 'Discount Rate:')
    ws3.cell(2, 2, 0.10)
    ws3.sheet_state = 'hidden'

    # ===== Named Ranges =====
    from openpyxl.workbook.defined_name import DefinedName

    # TaxRate
    tax_rate_ref = DefinedName('TaxRate', attr_text="Configuration!$B$1")
    wb.defined_names['TaxRate'] = tax_rate_ref

    # DiscountRate
    discount_rate_ref = DefinedName('DiscountRate', attr_text="Configuration!$B$2")
    wb.defined_names['DiscountRate'] = discount_rate_ref

    # ===== Workbook Properties =====
    wb.properties.title = 'Sample Sales Workbook'
    wb.properties.subject = 'Testing Excel Flattener'
    wb.properties.creator = 'Excel Flattener Test Suite'
    wb.properties.description = 'Sample workbook with various Excel features for testing'
    wb.properties.keywords = 'sales, test, sample'
    wb.properties.category = 'Test Data'
    wb.properties.company = 'Test Company Ltd'

    # Save
    output_path = Path('sample.xlsx')
    wb.save(output_path)

    print(f"✓ Created: {output_path}")
    print("\n⚠️  Note: This is an .xlsx file without VBA macros.")
    print("To create an .xlsm file with VBA:")
    print("1. Open sample.xlsx in Excel")
    print("2. Press Alt+F11 to open VBA editor")
    print("3. Insert > Module")
    print("4. Add sample VBA code (see below)")
    print("5. Save as sample.xlsm (Excel Macro-Enabled Workbook)")
    print("\nSample VBA code:")
    print("-" * 50)
    print("""
Sub CalculateTotals()
    '
    ' CalculateTotals Macro
    ' Recalculates all totals in the workbook
    '
    Application.CalculateFullRebuild
    MsgBox "Totals recalculated!", vbInformation
End Sub

Function GetTaxRate() As Double
    '
    ' Returns the tax rate from Configuration sheet
    '
    GetTaxRate = Worksheets("Configuration").Range("B1").Value
End Function
    """)
    print("-" * 50)


if __name__ == '__main__':
    create_sample_workbook()
