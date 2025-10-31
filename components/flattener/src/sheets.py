"""
Sheet data extraction (formulas, literal values, computed values, formats).

Extracts cell data from worksheets in a deterministic, diff-friendly format.
"""
import logging
from pathlib import Path
from typing import Dict, Any, List, Optional

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

from .normalizer import (
    normalise_cell_value,
    normalise_line_endings,
    sort_rows_by_address
)

logger = logging.getLogger(__name__)


class SheetExtractor:
    """
    Extracts data from a single worksheet.
    """

    def __init__(self, ws: Worksheet, include_computed: bool = False):
        """
        Initialise sheet extractor.

        Args:
            ws: Worksheet to extract
            include_computed: Whether to extract computed values
        """
        self.ws = ws
        self.include_computed = include_computed
        self.sheet_name = ws.title

        logger.debug(f"Initialised SheetExtractor for: {self.sheet_name}")

    def extract_formulas(self) -> List[Dict[str, str]]:
        """
        Extract all formulas from the sheet.

        Returns:
            List of {'address': 'A1', 'formula': '=SUM(A1:A10)'}
        """
        logger.debug(f"Extracting formulas from: {self.sheet_name}")

        formulas = []

        for row in self.ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # Formula
                    formula_value = cell.value
                    if formula_value:
                        # Extract formula text
                        # - If string: use as-is
                        # - If ArrayFormula object: use .text attribute
                        # - Otherwise: convert to string
                        if isinstance(formula_value, str):
                            formula_str = formula_value
                        elif hasattr(formula_value, 'text'):
                            formula_str = formula_value.text
                        else:
                            formula_str = str(formula_value)

                        # Ensure leading =
                        if formula_str and not formula_str.startswith('='):
                            formula_str = f'={formula_str}'

                        formulas.append({
                            'address': cell.coordinate,
                            'formula': formula_str
                        })

        # Sort by address
        formulas = sort_rows_by_address(formulas)

        logger.info(f"✓ Extracted {len(formulas)} formulas from {self.sheet_name}")
        return formulas

    def extract_literal_values(self) -> List[Dict[str, str]]:
        """
        Extract all literal (hardcoded) values from the sheet.

        Excludes:
        - Formulas
        - Empty cells
        - Computed values (if include_computed=False)

        Returns:
            List of {'address': 'A1', 'value': '42', 'type': 'number'}
        """
        logger.debug(f"Extracting literal values from: {self.sheet_name}")

        values = []

        for row in self.ws.iter_rows():
            for cell in row:
                # Skip formulas
                if cell.data_type == 'f':
                    continue

                # Skip empty
                if cell.value is None or cell.value == '':
                    continue

                # Extract value with type
                normalised_value = normalise_cell_value(cell.value)
                cell_type = self._get_cell_type(cell)

                values.append({
                    'address': cell.coordinate,
                    'value': normalised_value,
                    'type': cell_type
                })

        # Sort by address
        values = sort_rows_by_address(values)

        logger.info(f"✓ Extracted {len(values)} literal values from {self.sheet_name}")
        return values

    def extract_computed_values(self) -> List[Dict[str, str]]:
        """
        Extract computed values (results of formulas).

        Only called if include_computed=True.

        Returns:
            List of {'address': 'A1', 'value': '42', 'type': 'number'}
        """
        if not self.include_computed:
            return []

        logger.debug(f"Extracting computed values from: {self.sheet_name}")

        values = []

        for row in self.ws.iter_rows():
            for cell in row:
                # Only process formulas
                if cell.data_type != 'f':
                    continue

                # Get cached value (computed result)
                cached_value = cell.value
                if cached_value is None or cached_value == '':
                    continue

                # Extract value with type
                normalised_value = normalise_cell_value(cached_value)
                cell_type = self._get_cell_type(cell)

                values.append({
                    'address': cell.coordinate,
                    'value': normalised_value,
                    'type': cell_type
                })

        # Sort by address
        values = sort_rows_by_address(values)

        logger.info(f"✓ Extracted {len(values)} computed values from {self.sheet_name}")
        return values

    def extract_formats(self) -> List[Dict[str, Any]]:
        """
        Extract cell formatting information.

        Includes:
        - Number format
        - Font (name, size, bold, italic, colour)
        - Fill (background colour)
        - Border
        - Alignment

        Returns:
            List of {'address': 'A1', 'format': {...}}
        """
        logger.debug(f"Extracting formats from: {self.sheet_name}")

        formats = []

        for row in self.ws.iter_rows():
            for cell in row:
                # Skip cells with default formatting
                format_info = self._extract_cell_format(cell)
                if format_info:
                    formats.append({
                        'address': cell.coordinate,
                        'format': format_info
                    })

        # Sort by address
        formats = sort_rows_by_address(formats)

        logger.info(f"✓ Extracted {len(formats)} formatted cells from {self.sheet_name}")
        return formats

    def _get_cell_type(self, cell: Cell) -> str:
        """
        Determine cell type.

        Args:
            cell: Cell object

        Returns:
            Type string: number, text, boolean, date, error
        """
        if cell.is_date:
            return 'date'
        elif cell.data_type == 'n':
            return 'number'
        elif cell.data_type == 'b':
            return 'boolean'
        elif cell.data_type == 'e':
            return 'error'
        elif cell.data_type == 's':
            return 'text'
        else:
            return 'text'

    def _extract_cell_format(self, cell: Cell) -> Optional[Dict[str, Any]]:
        """
        Extract formatting information from a cell.

        Args:
            cell: Cell object

        Returns:
            Dictionary of format properties or None if default
        """
        format_info = {}

        try:
            # Number format
            if cell.number_format and cell.number_format != 'General':
                format_info['number_format'] = cell.number_format

            # Font
            if cell.font:
                font = cell.font
                font_info = {}
                if font.name:
                    font_info['name'] = font.name
                if font.size:
                    font_info['size'] = font.size
                if font.bold:
                    font_info['bold'] = True
                if font.italic:
                    font_info['italic'] = True
                if font.underline and font.underline != 'none':
                    font_info['underline'] = font.underline
                if font.color and hasattr(font.color, 'rgb') and font.color.rgb:
                    font_info['colour'] = font.color.rgb

                if font_info:
                    format_info['font'] = font_info

            # Fill (background)
            if cell.fill and cell.fill.fill_type:
                fill = cell.fill
                if fill.fill_type != 'none':
                    fill_info = {'type': fill.fill_type}
                    if hasattr(fill, 'fgColor') and fill.fgColor and hasattr(fill.fgColor, 'rgb') and fill.fgColor.rgb:
                        fill_info['colour'] = fill.fgColor.rgb
                    format_info['fill'] = fill_info

            # Border
            if cell.border:
                border = cell.border
                border_info = {}
                for side in ['left', 'right', 'top', 'bottom']:
                    side_obj = getattr(border, side, None)
                    if side_obj and side_obj.style:
                        border_info[side] = side_obj.style
                if border_info:
                    format_info['border'] = border_info

            # Alignment
            if cell.alignment:
                alignment = cell.alignment
                alignment_info = {}
                if alignment.horizontal:
                    alignment_info['horizontal'] = alignment.horizontal
                if alignment.vertical:
                    alignment_info['vertical'] = alignment.vertical
                if alignment.wrap_text:
                    alignment_info['wrap_text'] = True
                if alignment_info:
                    format_info['alignment'] = alignment_info

        except Exception as e:
            logger.warning(f"Error extracting format for {cell.coordinate}: {e}")

        return format_info if format_info else None


def write_formulas_file(sheet_name: str, formulas: List[Dict[str, str]], output_path: Path, sort_order: str = 'row') -> None:
    """
    Write formulas to text file.

    Format: A1: =SUM(A1:A10)

    Args:
        sheet_name: Name of sheet
        formulas: List of formula dictionaries
        output_path: Path to output file
        sort_order: 'row' for row-major (A1,A2,A3,B1...) or 'column' for column-major (A1,B1,C1,A2...)
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Apply appropriate sorting
    if sort_order == 'column':
        from .normalizer import sort_columns_by_address
        sorted_formulas = sort_columns_by_address(formulas)
        order_desc = 'column-by-column (A1, B1, C1, A2, B2, C2...)'
    else:
        # formulas are already sorted by row-major order from extract_formulas()
        sorted_formulas = formulas
        order_desc = 'row-by-row (A1, A2, A3, B1, B2, B3...)'

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(f'# Formulas: {sheet_name}\n')
        f.write(f'# Order: {order_desc}\n')
        f.write('# ' + '=' * 50 + '\n\n')

        for item in sorted_formulas:
            formula = normalise_line_endings(item['formula'])
            f.write(f"{item['address']}: {formula}\n")

    logger.debug(f"Wrote formulas ({sort_order}-order) to: {output_path}")


def write_values_file(sheet_name: str, values: List[Dict[str, str]], output_path: Path, file_type: str = 'literal') -> None:
    """
    Write values to text file.

    Format: A1: 42 (number)

    Args:
        sheet_name: Name of sheet
        values: List of value dictionaries
        output_path: Path to output file
        file_type: 'literal' or 'computed'
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    title = 'Literal Values' if file_type == 'literal' else 'Computed Values'

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(f'# {title}: {sheet_name}\n')
        f.write('# ' + '=' * 50 + '\n\n')

        for item in values:
            value = normalise_line_endings(item['value'])
            f.write(f"{item['address']}: {value} ({item['type']})\n")

    logger.debug(f"Wrote {file_type} values to: {output_path}")


def write_formats_file(sheet_name: str, formats: List[Dict[str, Any]], output_path: Path) -> None:
    """
    Write formatting information to text file.

    Format:
    A1:
      number_format: 0.00
      font:
        name: Arial
        size: 12
        bold: true

    Args:
        sheet_name: Name of sheet
        formats: List of format dictionaries
        output_path: Path to output file
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(f'# Formats: {sheet_name}\n')
        f.write('# ' + '=' * 50 + '\n\n')

        for item in formats:
            f.write(f"{item['address']}:\n")
            _write_format_dict(f, item['format'], indent=2)
            f.write('\n')

    logger.debug(f"Wrote formats to: {output_path}")


def _write_format_dict(f, format_dict: Dict[str, Any], indent: int = 0) -> None:
    """
    Recursively write format dictionary with indentation.

    Args:
        f: File handle
        format_dict: Dictionary to write
        indent: Number of spaces to indent
    """
    for key, value in format_dict.items():
        if isinstance(value, dict):
            f.write(f"{' ' * indent}{key}:\n")
            _write_format_dict(f, value, indent + 2)
        else:
            f.write(f"{' ' * indent}{key}: {value}\n")
