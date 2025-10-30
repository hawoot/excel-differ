"""
Main Flattener class - orchestrates Excel workbook flattening.

Coordinates all extraction modules to produce a deterministic, diff-friendly
text representation of an Excel workbook.
"""
import logging
import signal
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional, List

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from .utils import (
    get_file_hash,
    create_flat_root_name,
    setup_logging
)
from .manifest import Manifest
from .metadata import extract_metadata, write_metadata_file
from .structure import extract_structure, write_structure_file
from .sheets import SheetExtractor, write_formulas_file, write_values_file, write_formats_file
from .vba import extract_vba, write_vba_files, write_vba_summary
from .tables import extract_tables, write_tables_file, extract_autofilters, write_autofilters_file
from .charts import extract_charts, write_charts_file
from .named_ranges import extract_named_ranges, write_named_ranges_file

logger = logging.getLogger(__name__)


class TimeoutError(Exception):
    """Raised when extraction exceeds timeout."""
    pass


class Flattener:
    """
    Excel workbook flattener.

    Converts Excel workbooks to deterministic text representations.
    """

    def __init__(
        self,
        output_dir: Path,
        include_computed: bool = False,
        timeout: int = 900,
        max_file_size_mb: int = 200
    ):
        """
        Initialise flattener.

        Args:
            output_dir: Directory to write flat outputs
            include_computed: Whether to extract computed values
            timeout: Maximum extraction time in seconds
            max_file_size_mb: Maximum file size in MB
        """
        self.output_dir = Path(output_dir)
        self.include_computed = include_computed
        self.timeout = timeout
        self.max_file_size_mb = max_file_size_mb

        self.output_dir.mkdir(parents=True, exist_ok=True)

        logger.info(f"Flattener initialised (output: {output_dir}, "
                   f"computed: {include_computed}, timeout: {timeout}s)")

    def flatten(
        self,
        excel_file: Path,
        origin_repo: Optional[str] = None,
        origin_path: Optional[str] = None,
        origin_commit: Optional[str] = None,
        origin_commit_message: Optional[str] = None
    ) -> Path:
        """
        Flatten an Excel workbook.

        Args:
            excel_file: Path to Excel file
            origin_repo: Git repository URL (optional)
            origin_path: Path in repository (optional)
            origin_commit: Git commit SHA (optional)
            origin_commit_message: Commit message (optional)

        Returns:
            Path to flat root directory

        Raises:
            ValueError: If file is invalid or too large
            TimeoutError: If extraction exceeds timeout
            Exception: For other extraction errors
        """
        # Validate file
        self._validate_file(excel_file)

        # Setup timeout handler
        if self.timeout > 0:
            signal.signal(signal.SIGALRM, self._timeout_handler)
            signal.alarm(self.timeout)

        try:
            logger.info(f"=" * 70)
            logger.info(f"Starting extraction: {excel_file.name}")
            logger.info(f"=" * 70)

            # Calculate file hash
            file_hash = get_file_hash(excel_file)
            logger.info(f"File hash: {file_hash[:16]}...")

            # Create flat root directory
            timestamp = datetime.now(timezone.utc)
            flat_root_name = create_flat_root_name(excel_file.stem, timestamp, file_hash[:8])
            flat_root = self.output_dir / flat_root_name
            flat_root.mkdir(parents=True, exist_ok=True)

            logger.info(f"Flat root: {flat_root}")

            # Initialise manifest
            manifest = Manifest(
                workbook_filename=excel_file.name,
                original_sha256=file_hash,
                include_computed=self.include_computed
            )

            # Set origin if provided
            if origin_repo or origin_path or origin_commit:
                manifest.set_origin(
                    origin_repo=origin_repo,
                    origin_path=origin_path,
                    origin_commit=origin_commit,
                    origin_commit_message=origin_commit_message
                )

            # Load workbook
            logger.info("Loading workbook...")
            wb = self._load_workbook(excel_file)

            # Extract all components
            self._extract_metadata(wb, flat_root, manifest)
            self._extract_structure(wb, flat_root, manifest)
            self._extract_sheets(wb, flat_root, manifest)
            self._extract_vba(excel_file, flat_root, manifest)
            self._extract_tables(wb, flat_root, manifest)
            self._extract_charts(wb, flat_root, manifest)
            self._extract_named_ranges(wb, flat_root, manifest)

            # Save manifest
            manifest_path = flat_root / 'manifest.json'
            manifest.save(manifest_path)
            manifest.add_file(manifest_path, flat_root)

            logger.info(f"=" * 70)
            logger.info(f"✓ Extraction complete: {flat_root_name}")
            logger.info(f"  Total files: {len(manifest.files)}")
            logger.info(f"  Warnings: {len(manifest.warnings)}")
            logger.info(f"=" * 70)

            return flat_root

        finally:
            # Cancel timeout
            if self.timeout > 0:
                signal.alarm(0)

    def _validate_file(self, file_path: Path) -> None:
        """
        Validate Excel file.

        Args:
            file_path: Path to file

        Raises:
            ValueError: If file is invalid
        """
        if not file_path.exists():
            raise ValueError(f"File not found: {file_path}")

        if not file_path.is_file():
            raise ValueError(f"Not a file: {file_path}")

        # Check file size
        file_size_mb = file_path.stat().st_size / (1024 * 1024)
        if file_size_mb > self.max_file_size_mb:
            raise ValueError(
                f"File too large: {file_size_mb:.1f}MB "
                f"(max: {self.max_file_size_mb}MB)"
            )

        # Check extension
        valid_extensions = ['.xlsx', '.xlsm', '.xlsb', '.xls']
        if file_path.suffix.lower() not in valid_extensions:
            raise ValueError(
                f"Unsupported file type: {file_path.suffix} "
                f"(supported: {', '.join(valid_extensions)})"
            )

        logger.info(f"✓ File validated ({file_size_mb:.1f}MB)")

    def _load_workbook(self, file_path: Path) -> Workbook:
        """
        Load workbook with openpyxl.

        Args:
            file_path: Path to Excel file

        Returns:
            Workbook object
        """
        try:
            # Load with data_only=False to get formulas
            wb = load_workbook(
                filename=str(file_path),
                data_only=False,
                keep_vba=False  # VBA extracted separately with oletools
            )
            logger.info(f"✓ Workbook loaded ({len(wb.worksheets)} sheets)")
            return wb

        except Exception as e:
            logger.error(f"Failed to load workbook: {e}", exc_info=True)
            raise

    def _extract_metadata(self, wb: Workbook, flat_root: Path, manifest: Manifest) -> None:
        """Extract workbook metadata."""
        logger.info("Extracting metadata...")
        try:
            metadata = extract_metadata(wb)
            metadata_path = flat_root / 'metadata.txt'
            write_metadata_file(metadata, metadata_path)
            manifest.add_file(metadata_path, flat_root)
        except Exception as e:
            logger.error(f"Error extracting metadata: {e}", exc_info=True)
            manifest.add_warning(f"Metadata extraction failed: {e}")

    def _extract_structure(self, wb: Workbook, flat_root: Path, manifest: Manifest) -> None:
        """Extract workbook structure."""
        logger.info("Extracting structure...")
        try:
            structure = extract_structure(wb)

            # Add sheets to manifest
            for sheet_info in structure:
                manifest.add_sheet(
                    index=sheet_info['index'],
                    name=sheet_info['name'],
                    sheetId=sheet_info['sheetId'],
                    visible=sheet_info['visible']
                )

            structure_path = flat_root / 'structure.txt'
            write_structure_file(structure, structure_path)
            manifest.add_file(structure_path, flat_root)

        except Exception as e:
            logger.error(f"Error extracting structure: {e}", exc_info=True)
            manifest.add_warning(f"Structure extraction failed: {e}")

    def _extract_sheets(self, wb: Workbook, flat_root: Path, manifest: Manifest) -> None:
        """Extract sheet data."""
        logger.info("Extracting sheets...")

        sheets_dir = flat_root / 'sheets'
        sheets_dir.mkdir(exist_ok=True)

        for ws in wb.worksheets:
            sheet_name = ws.title
            logger.info(f"  Processing sheet: {sheet_name}")

            try:
                extractor = SheetExtractor(ws, include_computed=self.include_computed)

                # Create sheet directory
                # Sanitise sheet name for filesystem
                safe_name = self._sanitise_sheet_name(sheet_name)
                sheet_dir = sheets_dir / safe_name
                sheet_dir.mkdir(exist_ok=True)

                # Extract formulas
                formulas = extractor.extract_formulas()
                if formulas:
                    formulas_path = sheet_dir / 'formulas.txt'
                    write_formulas_file(sheet_name, formulas, formulas_path)
                    manifest.add_file(formulas_path, flat_root)

                # Extract literal values
                literal_values = extractor.extract_literal_values()
                if literal_values:
                    literal_path = sheet_dir / 'literal-values.txt'
                    write_values_file(sheet_name, literal_values, literal_path, file_type='literal')
                    manifest.add_file(literal_path, flat_root)

                # Extract computed values (if enabled)
                if self.include_computed:
                    computed_values = extractor.extract_computed_values()
                    if computed_values:
                        computed_path = sheet_dir / 'computed-values.txt'
                        write_values_file(sheet_name, computed_values, computed_path, file_type='computed')
                        manifest.add_file(computed_path, flat_root)

                # Extract formats
                formats = extractor.extract_formats()
                if formats:
                    formats_path = sheet_dir / 'formats.txt'
                    write_formats_file(sheet_name, formats, formats_path)
                    manifest.add_file(formats_path, flat_root)

            except Exception as e:
                logger.error(f"Error extracting sheet {sheet_name}: {e}", exc_info=True)
                manifest.add_warning(f"Sheet '{sheet_name}' extraction failed: {e}")

    def _extract_vba(self, excel_file: Path, flat_root: Path, manifest: Manifest) -> None:
        """Extract VBA macros."""
        logger.info("Extracting VBA...")
        try:
            vba_info = extract_vba(excel_file)

            if vba_info:
                vba_dir = flat_root / 'vba'
                vba_dir.mkdir(exist_ok=True)

                # Write VBA files
                vba_files = write_vba_files(vba_info, vba_dir)
                for vba_file in vba_files:
                    manifest.add_file(vba_file, flat_root)

                # Write summary
                summary_path = vba_dir / 'vba-summary.txt'
                write_vba_summary(vba_info, summary_path)
                manifest.add_file(summary_path, flat_root)

        except Exception as e:
            logger.error(f"Error extracting VBA: {e}", exc_info=True)
            manifest.add_warning(f"VBA extraction failed: {e}")

    def _extract_tables(self, wb: Workbook, flat_root: Path, manifest: Manifest) -> None:
        """Extract tables and autofilters."""
        logger.info("Extracting tables...")
        try:
            # Tables
            tables = extract_tables(wb)
            if tables:
                tables_path = flat_root / 'tables.txt'
                write_tables_file(tables, tables_path)
                manifest.add_file(tables_path, flat_root)

            # Autofilters
            autofilters = extract_autofilters(wb)
            if autofilters:
                autofilters_path = flat_root / 'autofilters.txt'
                write_autofilters_file(autofilters, autofilters_path)
                manifest.add_file(autofilters_path, flat_root)

        except Exception as e:
            logger.error(f"Error extracting tables: {e}", exc_info=True)
            manifest.add_warning(f"Tables extraction failed: {e}")

    def _extract_charts(self, wb: Workbook, flat_root: Path, manifest: Manifest) -> None:
        """Extract charts."""
        logger.info("Extracting charts...")
        try:
            charts = extract_charts(wb)
            if charts:
                charts_path = flat_root / 'charts.txt'
                write_charts_file(charts, charts_path)
                manifest.add_file(charts_path, flat_root)

        except Exception as e:
            logger.error(f"Error extracting charts: {e}", exc_info=True)
            manifest.add_warning(f"Charts extraction failed: {e}")

    def _extract_named_ranges(self, wb: Workbook, flat_root: Path, manifest: Manifest) -> None:
        """Extract named ranges."""
        logger.info("Extracting named ranges...")
        try:
            named_ranges = extract_named_ranges(wb)
            if named_ranges:
                named_ranges_path = flat_root / 'named-ranges.txt'
                write_named_ranges_file(named_ranges, named_ranges_path)
                manifest.add_file(named_ranges_path, flat_root)

        except Exception as e:
            logger.error(f"Error extracting named ranges: {e}", exc_info=True)
            manifest.add_warning(f"Named ranges extraction failed: {e}")

    def _sanitise_sheet_name(self, name: str) -> str:
        """
        Sanitise sheet name for filesystem.

        Args:
            name: Sheet name

        Returns:
            Sanitised name
        """
        # Replace problematic characters
        replacements = {
            '/': '_',
            '\\': '_',
            ':': '_',
            '*': '_',
            '?': '_',
            '"': '_',
            '<': '_',
            '>': '_',
            '|': '_',
        }

        sanitised = name
        for char, replacement in replacements.items():
            sanitised = sanitised.replace(char, replacement)

        # Remove leading/trailing spaces and dots
        sanitised = sanitised.strip().strip('.')

        # Ensure not empty
        if not sanitised:
            sanitised = 'sheet'

        return sanitised

    def _timeout_handler(self, signum, frame):
        """Handle timeout signal."""
        raise TimeoutError(f"Extraction exceeded timeout of {self.timeout}s")
