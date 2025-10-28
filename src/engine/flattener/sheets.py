"""
Sheet-level data extraction.
Extracts formulas, values, cell formats, merged ranges, comments, etc.
"""
from pathlib import Path
from typing import Dict, Any, Optional
import logging
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell

from src.engine.flattener.normalizer import (
    normalize_formula,
    normalize_cell_value,
    normalize_color,
    sort_key_for_cell_address,
    normalize_sheet_name,
)
from src.engine.flattener.manifest import Manifest

logger = logging.getLogger(__name__)


def extract_sheet_data(
    sheet: Worksheet,
    sheet_index: int,
    sheets_dir: Path,
    include_evaluated: bool,
    manifest: Manifest,
) -> None:
    """
    Extract all data from a worksheet.

    Args:
        sheet: Worksheet object
        sheet_index: Sheet index (1-based)
        sheets_dir: Directory to write sheet files
        include_evaluated: Whether to include evaluated values
        manifest: Manifest object to add warnings
    """
    # Create file prefix
    safe_name = normalize_sheet_name(sheet.title)
    prefix = f"{sheet_index:02d}.{safe_name}"

    logger.debug(f"Extracting sheet: {sheet.title} ({prefix})")

    # Extract metadata
    _extract_sheet_metadata(sheet, sheet_index, sheets_dir, prefix, manifest)

    # Extract formulas and values
    _extract_formulas_and_values(sheet, sheets_dir, prefix, include_evaluated, manifest)

    # Extract cell formats
    _extract_cell_formats(sheet, sheets_dir, prefix, manifest)

    # Extract merged ranges
    _extract_merged_ranges(sheet, sheets_dir, prefix, manifest)

    # Extract data validations
    _extract_data_validations(sheet, sheets_dir, prefix, manifest)

    # Extract comments
    _extract_comments(sheet, sheets_dir, prefix, manifest)


def _extract_sheet_metadata(
    sheet: Worksheet,
    sheet_index: int,
    sheets_dir: Path,
    prefix: str,
    manifest: Manifest,
) -> None:
    """Extract sheet-level metadata (protection, tab color, etc.)."""
    import json

    metadata = {
        "sheetId": sheet.sheet_properties.sheetId if hasattr(sheet.sheet_properties, 'sheetId') else sheet_index,
        "visible": sheet.sheet_state == "visible",
        "state": sheet.sheet_state,
    }

    # Tab color
    if sheet.sheet_properties.tabColor:
        tab_color = sheet.sheet_properties.tabColor
        if hasattr(tab_color, 'rgb'):
            metadata["tab_color"] = f"#{tab_color.rgb}"
        elif hasattr(tab_color, 'theme'):
            metadata["tab_color"] = f"theme:{tab_color.theme}"

    # Sheet protection
    if sheet.protection and sheet.protection.sheet:
        metadata["protection"] = {
            "sheet_protected": True,
            "password": bool(sheet.protection.password),
        }

    # Write metadata
    metadata_path = sheets_dir / f"{prefix}.metadata.json"
    with open(metadata_path, "w", encoding="utf-8") as f:
        json.dump(metadata, f, indent=2)


def _extract_formulas_and_values(
    sheet: Worksheet,
    sheets_dir: Path,
    prefix: str,
    include_evaluated: bool,
    manifest: Manifest,
) -> None:
    """Extract formulas and values from all cells."""
    formulas_file = sheets_dir / f"{prefix}.formulas.txt"
    values_hardcoded_file = sheets_dir / f"{prefix}.values_hardcoded.txt"
    values_evaluated_file = sheets_dir / f"{prefix}.values_evaluated.txt"

    formulas = []
    values_hardcoded = []
    values_evaluated = []

    # Iterate through all cells
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue

            cell_addr = cell.coordinate

            # Check if cell has a formula
            if cell.data_type == 'f':
                # Cell has a formula
                formula_text = str(cell.value)
                normalized_formula = normalize_formula(formula_text)
                formulas.append((cell_addr, normalized_formula))

                # For evaluated values, get the cached result
                if include_evaluated:
                    # In data_only=False mode, we can't get evaluated values
                    # We'd need to load the workbook again with data_only=True
                    # For now, note that this is cached
                    evaluated_value = _get_cached_formula_value(cell)
                    if evaluated_value is not None:
                        values_evaluated.append((cell_addr, f"{evaluated_value}|cached"))

            else:
                # Cell has a hardcoded value (not a formula)
                value_str = _format_cell_value(cell)
                if value_str:
                    values_hardcoded.append((cell_addr, value_str))

                # For evaluated values, it's the same as hardcoded for non-formula cells
                if include_evaluated and value_str:
                    values_evaluated.append((cell_addr, value_str))

    # Sort all lists by cell address (row-major order)
    formulas.sort(key=lambda x: sort_key_for_cell_address(x[0]))
    values_hardcoded.sort(key=lambda x: sort_key_for_cell_address(x[0]))
    values_evaluated.sort(key=lambda x: sort_key_for_cell_address(x[0]))

    # Write formulas file
    with open(formulas_file, "w", encoding="utf-8") as f:
        f.write("# Formulas\n")
        f.write("# ADDRESS\tFORMULA\n\n")
        for addr, formula in formulas:
            f.write(f"{addr}\t{formula}\n")

    # Write values_hardcoded file (MANDATORY)
    with open(values_hardcoded_file, "w", encoding="utf-8") as f:
        f.write("# Hard-coded Values (non-formula cells only)\n")
        f.write("# ADDRESS\tVALUE\n\n")
        for addr, value in values_hardcoded:
            f.write(f"{addr}\t{value}\n")

    # Write values_evaluated file (OPTIONAL)
    if include_evaluated:
        with open(values_evaluated_file, "w", encoding="utf-8") as f:
            f.write("# Evaluated Values (all cells, including formula results)\n")
            f.write("# ADDRESS\tVALUE\n\n")
            for addr, value in values_evaluated:
                f.write(f"{addr}\t{value}\n")


def _extract_cell_formats(
    sheet: Worksheet,
    sheets_dir: Path,
    prefix: str,
    manifest: Manifest,
) -> None:
    """Extract cell formatting (number format, font, fill, alignment)."""
    formats_file = sheets_dir / f"{prefix}.cell_formats.txt"

    formats = []

    for row in sheet.iter_rows():
        for cell in row:
            # Skip cells with no formatting
            if cell.value is None and not _has_formatting(cell):
                continue

            cell_addr = cell.coordinate
            format_parts = []

            # Number format
            if cell.number_format and cell.number_format != "General":
                format_parts.append(f"number_format:{cell.number_format}")

            # Font
            if cell.font:
                font_parts = []
                if cell.font.name:
                    font_parts.append(f"name={cell.font.name}")
                if cell.font.size:
                    font_parts.append(f"size={cell.font.size}")
                if cell.font.bold:
                    font_parts.append("bold")
                if cell.font.italic:
                    font_parts.append("italic")
                if cell.font.color:
                    color = _format_color(cell.font.color)
                    if color:
                        font_parts.append(f"color={color}")

                if font_parts:
                    format_parts.append(f"font:{','.join(font_parts)}")

            # Fill
            if cell.fill and cell.fill.patternType:
                fill_parts = [f"pattern={cell.fill.patternType}"]
                if cell.fill.fgColor:
                    fg_color = _format_color(cell.fill.fgColor)
                    if fg_color:
                        fill_parts.append(f"fgColor={fg_color}")
                if cell.fill.bgColor:
                    bg_color = _format_color(cell.fill.bgColor)
                    if bg_color:
                        fill_parts.append(f"bgColor={bg_color}")

                format_parts.append(f"fill:{','.join(fill_parts)}")

            # Alignment
            if cell.alignment:
                align_parts = []
                if cell.alignment.horizontal:
                    align_parts.append(f"h={cell.alignment.horizontal}")
                if cell.alignment.vertical:
                    align_parts.append(f"v={cell.alignment.vertical}")
                if cell.alignment.wrap_text:
                    align_parts.append("wrap")

                if align_parts:
                    format_parts.append(f"align:{','.join(align_parts)}")

            # Border
            if cell.border and _has_borders(cell.border):
                format_parts.append("border:yes")

            if format_parts:
                formats.append((cell_addr, "|".join(format_parts)))

    # Sort by cell address
    formats.sort(key=lambda x: sort_key_for_cell_address(x[0]))

    # Write formats file
    with open(formats_file, "w", encoding="utf-8") as f:
        f.write("# Cell Formats\n")
        f.write("# ADDRESS\tFORMAT\n\n")
        for addr, format_str in formats:
            f.write(f"{addr}\t{format_str}\n")


def _extract_merged_ranges(
    sheet: Worksheet,
    sheets_dir: Path,
    prefix: str,
    manifest: Manifest,
) -> None:
    """Extract merged cell ranges."""
    merged_file = sheets_dir / f"{prefix}.merged_ranges.txt"

    merged_ranges = sorted(str(r) for r in sheet.merged_cells.ranges)

    with open(merged_file, "w", encoding="utf-8") as f:
        f.write("# Merged Ranges\n\n")
        for range_str in merged_ranges:
            f.write(f"{range_str}\n")


def _extract_data_validations(
    sheet: Worksheet,
    sheets_dir: Path,
    prefix: str,
    manifest: Manifest,
) -> None:
    """Extract data validation rules."""
    validations_file = sheets_dir / f"{prefix}.data_validations.txt"

    with open(validations_file, "w", encoding="utf-8") as f:
        f.write("# Data Validations\n")
        f.write("# RANGE\tTYPE\tFORMULA\n\n")

        for dv in sheet.data_validations.dataValidation:
            ranges = str(dv.sqref) if hasattr(dv, 'sqref') else ""
            val_type = dv.type if hasattr(dv, 'type') else ""
            formula1 = dv.formula1 if hasattr(dv, 'formula1') else ""

            f.write(f"{ranges}\t{val_type}\t{formula1}\n")


def _extract_comments(
    sheet: Worksheet,
    sheets_dir: Path,
    prefix: str,
    manifest: Manifest,
) -> None:
    """Extract cell comments."""
    comments_file = sheets_dir / f"{prefix}.comments.txt"

    comments = []

    for row in sheet.iter_rows():
        for cell in row:
            if cell.comment:
                cell_addr = cell.coordinate
                author = cell.comment.author or "Unknown"
                text = cell.comment.text or ""

                # Clean up text
                text = text.replace("\n", "\\n").replace("\r", "")

                comments.append((cell_addr, f"{author}|{text}"))

    # Sort by cell address
    comments.sort(key=lambda x: sort_key_for_cell_address(x[0]))

    with open(comments_file, "w", encoding="utf-8") as f:
        f.write("# Comments\n")
        f.write("# ADDRESS\tAUTHOR|TEXT\n\n")
        for addr, comment in comments:
            f.write(f"{addr}\t{comment}\n")


# Helper functions

def _format_cell_value(cell: Cell) -> str:
    """Format a cell's value for output."""
    if cell.value is None:
        return ""

    value_type = cell.data_type

    if value_type == 'b':
        # Boolean
        return "TRUE" if cell.value else "FALSE"
    elif value_type == 'n':
        # Number
        from src.engine.flattener.normalizer import normalize_number
        return normalize_number(cell.value)
    elif value_type == 'd':
        # Date
        from src.engine.flattener.normalizer import normalize_date
        return normalize_date(cell.value, number_format=cell.number_format)
    elif value_type == 's' or value_type == 'str':
        # String
        from src.engine.flattener.normalizer import normalize_string
        return normalize_string(str(cell.value))
    else:
        return str(cell.value)


def _get_cached_formula_value(cell: Cell) -> Optional[str]:
    """Get the cached value of a formula cell."""
    # In openpyxl, when data_only=False, cell.value contains the formula
    # The cached value might be in cell._value
    # This is a limitation - we'd need to reload with data_only=True
    # For now, return None
    return None


def _has_formatting(cell: Cell) -> bool:
    """Check if a cell has any non-default formatting."""
    if cell.font and (cell.font.bold or cell.font.italic or cell.font.color):
        return True
    if cell.fill and cell.fill.patternType != "none":
        return True
    if cell.border and _has_borders(cell.border):
        return True
    if cell.alignment and (cell.alignment.wrap_text or cell.alignment.horizontal):
        return True
    if cell.number_format and cell.number_format != "General":
        return True
    return False


def _has_borders(border) -> bool:
    """Check if a border has any styling."""
    if not border:
        return False
    return any([
        border.left and border.left.style,
        border.right and border.right.style,
        border.top and border.top.style,
        border.bottom and border.bottom.style,
    ])


def _format_color(color) -> Optional[str]:
    """Format an openpyxl color object."""
    if not color:
        return None

    if hasattr(color, 'rgb') and color.rgb:
        # RGB color
        rgb = color.rgb
        if isinstance(rgb, str) and len(rgb) >= 6:
            # Remove alpha channel if present (AARRGGBB -> RRGGBB)
            if len(rgb) == 8:
                rgb = rgb[2:]
            return f"#{rgb}"

    if hasattr(color, 'theme') and color.theme is not None:
        return f"theme:{color.theme}"

    if hasattr(color, 'indexed') and color.indexed is not None:
        return f"indexed:{color.indexed}"

    return None
