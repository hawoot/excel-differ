#!/usr/bin/env python3
"""
Quick test to verify multiprocessing backend works.
Run this while the API server is running.
"""
import requests
import time
import io
from openpyxl import Workbook

API_URL = "http://localhost:8000"

# Create a simple test Excel file
wb = Workbook()
ws = wb.active
ws['A1'] = 'Hello'
ws['B1'] = 'World'
ws['A2'] = 42
ws['B2'] = '=A2*2'

# Save to bytes
excel_bytes = io.BytesIO()
wb.save(excel_bytes)
excel_bytes.seek(0)

print("📤 Submitting flatten job...")
response = requests.post(
    f"{API_URL}/api/v1/flatten",
    files={"file": ("test.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    data={"format": "zip", "include_evaluated": "false"}
)

if response.status_code != 202:
    print(f"❌ Failed to submit job: {response.status_code}")
    print(response.text)
    exit(1)

job_id = response.json()["job_id"]
print(f"✅ Job submitted: {job_id}")

# Poll for completion
print("⏳ Polling for job completion...")
for i in range(30):  # Try for 30 seconds
    time.sleep(1)

    response = requests.get(f"{API_URL}/api/v1/jobs/{job_id}")
    job_data = response.json()

    status = job_data["status"]
    print(f"   [{i+1}s] Status: {status}")

    if status == "success":
        print("✅ Job completed successfully!")
        print(f"   Result keys: {list(job_data.get('result', {}).keys())}")
        exit(0)
    elif status == "failed":
        print(f"❌ Job failed: {job_data.get('error')}")
        exit(1)
    elif status not in ["queued", "running"]:
        print(f"⚠️  Unexpected status: {status}")
        exit(1)

print("⏰ Timeout waiting for job completion")
print("   The job is still running or stuck.")
print(f"   Check job file: /tmp/excel-differ/jobs/{job_id}.json")
exit(1)
